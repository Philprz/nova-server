"""
Module de parsing multi-format pour les fichiers de tarifs fournisseurs.
Supporte: PDF, Excel (.xlsx, .xls), CSV, images (via OCR)
"""

import os
import re
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import json

logger = logging.getLogger(__name__)

# Extensions supportées par type
SUPPORTED_EXTENSIONS = {
    'pdf': ['.pdf'],
    'excel': ['.xlsx', '.xls'],
    'csv': ['.csv'],
    'image': ['.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif']
}

ALL_SUPPORTED_EXTENSIONS = [ext for exts in SUPPORTED_EXTENSIONS.values() for ext in exts]


def get_file_type(file_path: str) -> Optional[str]:
    """Détermine le type de fichier basé sur l'extension."""
    ext = Path(file_path).suffix.lower()
    for file_type, extensions in SUPPORTED_EXTENSIONS.items():
        if ext in extensions:
            return file_type
    return None


def is_supported_file(file_path: str) -> bool:
    """Vérifie si le fichier est supporté."""
    return get_file_type(file_path) is not None


def detect_currency(text: str) -> str:
    """Détecte la devise dans une chaîne de texte.

    Retourne le code ISO : 'EUR', 'GBP', 'USD' ou 'CHF'.
    Priorité : GBP > USD > CHF > EUR (défaut).
    """
    if not text:
        return "EUR"
    t = str(text).upper()
    if '£' in t or 'GBP' in t or 'POUND' in t or 'STERLING' in t:
        return "GBP"
    if '$' in t or 'USD' in t or 'DOLLAR' in t:
        return "USD"
    if 'CHF' in t or 'FRANC SUISSE' in t:
        return "CHF"
    return "EUR"


def extract_price(text: str) -> Optional[float]:
    """Extrait un prix d'une chaîne de texte (valeur numérique seule).

    La devise associée est détectée séparément via detect_currency().
    """
    if not text:
        return None

    # Patterns ordonnés : avec symbole monétaire en premier (plus précis)
    patterns = [
        r'(\d+[.,]\d{2})\s*(?:€|EUR|euro)',   # 123.45 € ou 123,45 EUR
        r'(?:€|EUR)\s*(\d+[.,]\d{2})',         # € 123.45
        r'(\d+[.,]\d{2})\s*(?:£|GBP)',         # 123.45 £ ou GBP
        r'(?:£|GBP)\s*(\d+[.,]\d{2})',         # £ 123.45
        r'(\d+[.,]\d{2})\s*(?:\$|USD)',         # 123.45 $ ou USD
        r'(?:\$|USD)\s*(\d+[.,]\d{2})',         # $ 123.45
        r'(\d+[.,]\d{2})\s*(?:CHF)',            # 123.45 CHF
        r'CHF\s*(\d+[.,]\d{2})',                # CHF 123.45
        r'(\d+[.,]\d{2,3})',                    # Nombre décimal seul
        r'(\d+)\s*(?:€|EUR|euro|£|GBP|\$|USD|CHF)',  # Entier avec devise
    ]

    for pattern in patterns:
        match = re.search(pattern, str(text), re.IGNORECASE)
        if match:
            price_str = match.group(1).replace(',', '.')
            try:
                return float(price_str)
            except ValueError:
                continue
    return None


def extract_reference(text: str) -> Optional[str]:
    """Extrait une référence produit d'une chaîne."""
    if not text:
        return None

    # Patterns communs pour les références
    patterns = [
        r'(?:ref|réf|reference|référence)[.:\s]*([A-Z0-9\-_]+)',
        r'([A-Z]{2,4}[\-_]?\d{3,})',  # XX-1234 ou ABC1234
        r'(\d{6,})',                    # Code numérique 6+ chiffres
    ]

    for pattern in patterns:
        match = re.search(pattern, str(text), re.IGNORECASE)
        if match:
            return match.group(1).strip()
    return None


class PDFParser:
    """Parser pour les fichiers PDF."""

    @staticmethod
    def parse(file_path: str) -> List[Dict[str, Any]]:
        """Parse un fichier PDF et extrait les données de tarifs."""
        products = []

        try:
            import fitz  # PyMuPDF
        except ImportError:
            logger.warning("PyMuPDF non installé. Installation: pip install pymupdf")
            try:
                # Fallback sur pdfplumber
                import pdfplumber
                return PDFParser._parse_with_pdfplumber(file_path)
            except ImportError:
                logger.error("Aucune bibliothèque PDF disponible")
                return []

        try:
            doc = fitz.open(file_path)
            full_text = ""

            for page in doc:
                full_text += page.get_text()

            doc.close()

            # Extraction des données structurées
            products = PDFParser._extract_products_from_text(full_text, file_path)

        except Exception as e:
            logger.error(f"Erreur parsing PDF {file_path}: {e}")

        return products

    @staticmethod
    def _parse_with_pdfplumber(file_path: str) -> List[Dict[str, Any]]:
        """Parse PDF avec pdfplumber (fallback)."""
        import pdfplumber
        products = []

        try:
            with pdfplumber.open(file_path) as pdf:
                full_text = ""
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        full_text += text + "\n"

                    # Extraction des tableaux
                    tables = page.extract_tables()
                    for table in tables:
                        products.extend(PDFParser._parse_table(table, file_path))

                # Si pas de tableaux, extraire du texte
                if not products:
                    products = PDFParser._extract_products_from_text(full_text, file_path)

        except Exception as e:
            logger.error(f"Erreur pdfplumber {file_path}: {e}")

        return products

    @staticmethod
    def _parse_table(table: List[List], file_path: str) -> List[Dict[str, Any]]:
        """Parse un tableau extrait d'un PDF."""
        products = []
        if not table or len(table) < 2:
            return products

        # Première ligne = en-têtes
        headers = [str(h).lower() if h else '' for h in table[0]]

        # Mapping des colonnes
        col_mapping = {
            'reference': ['ref', 'réf', 'reference', 'référence', 'code', 'sku', 'article',
                          'item no', 'part no', 'no.', 'part number', 'item number'],
            'designation': ['designation', 'désignation', 'description', 'libellé', 'libelle', 'nom', 'produit'],
            'price': ['prix', 'price', 'tarif', 'pu', 'p.u.', 'montant', 'ht',
                      'selling', 'list price', 'net price', 'unit price', 'sell price'],
            'supplier': ['fournisseur', 'supplier', 'marque', 'brand'],
            'delivery': ['délai', 'delai', 'livraison', 'delivery', 'dispo']
        }

        col_indices = {}
        for field, keywords in col_mapping.items():
            for i, header in enumerate(headers):
                if any(kw in header for kw in keywords):
                    col_indices[field] = i
                    break

        # Détecter la devise depuis le header de la colonne prix (ex: "Prix £" → GBP)
        price_header = headers[col_indices['price']] if 'price' in col_indices else ''
        table_currency = detect_currency(price_header)

        # Parser les lignes de données
        for row in table[1:]:
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            # Détecter la devise dans la cellule prix (priorité sur header)
            raw_price_cell = str(row[col_indices['price']]) if 'price' in col_indices and col_indices['price'] < len(row) else ''
            cell_currency = detect_currency(raw_price_cell)
            row_currency = cell_currency if cell_currency != "EUR" else table_currency

            product = {
                'supplier_reference': row[col_indices.get('reference')] if 'reference' in col_indices and col_indices['reference'] < len(row) else None,
                'designation': row[col_indices.get('designation')] if 'designation' in col_indices and col_indices['designation'] < len(row) else None,
                'unit_price': extract_price(row[col_indices.get('price')]) if 'price' in col_indices and col_indices['price'] < len(row) else None,
                'supplier_name': row[col_indices.get('supplier')] if 'supplier' in col_indices and col_indices['supplier'] < len(row) else None,
                'delivery_time': row[col_indices.get('delivery')] if 'delivery' in col_indices and col_indices['delivery'] < len(row) else None,
                'currency': row_currency,
                'additional_data': {'source_file': Path(file_path).name}
            }

            # Ne garder que si on a au moins une référence ou une désignation
            if product['supplier_reference'] or product['designation']:
                products.append(product)

        return products

    @staticmethod
    def _extract_products_from_text(text: str, file_path: str) -> List[Dict[str, Any]]:
        """Extraction heuristique de produits depuis du texte brut."""
        products = []
        lines = text.split('\n')

        for line in lines:
            line = line.strip()
            if not line or len(line) < 10:
                continue

            # Chercher des patterns de produits
            ref = extract_reference(line)
            price = extract_price(line)

            if ref or price:
                product = {
                    'supplier_reference': ref,
                    'designation': line[:200] if len(line) > 200 else line,
                    'unit_price': price,
                    'currency': detect_currency(line),
                    'additional_data': {'source_file': Path(file_path).name, 'raw_line': line}
                }
                products.append(product)

        return products


class ExcelParser:
    """
    Parser robuste pour fichiers Excel (pièces jointes RFQ client).

    Stratégie en 2 passes :
      1. Détection de la ligne d'en-tête par keywords (multi-langue, étendue).
         Si ≥2 colonnes mappées → chemin structuré.
      2. Fallback heuristique si le mapping est insuffisant :
         lecture positionnelle colonne par colonne.
    Règle absolue : aucune ligne non-vide n'est supprimée.
    """

    # Keywords étendus, multi-langue (FR/EN/TR/IT/DE)
    # Note : 'no.', 'n°', 'nr', 'pos' retirés de 'reference' — ce sont des
    # compteurs de lignes (très présents dans les RFQ industriels) et non des
    # colonnes de référence article. Ils feraient masquer la vraie colonne
    # "Part Number" / "Reference" qui arrive en colonne 2 ou 3.
    _COL_KEYWORDS: Dict[str, List[str]] = {
        'reference': [
            'ref', 'réf', 'reference', 'référence', 'code', 'sku', 'article',
            'code article', 'code produit', 'item no', 'item no.', 'part no',
            'part no.', 'part number', 'item number', 'part', 'pn', 'partnumber',
            'malzeme', 'malzeme kodu', 'codice', 'articolo', 'artikel',
        ],
        'designation': [
            'designation', 'désignation', 'description', 'libellé', 'libelle',
            'nom', 'produit', 'intitulé', 'item description', 'item name',
            'item', 'name', 'article name', 'denominazione', 'bezeichnung',
            'tanim', 'aciklama', 'açıklama', 'malzeme tanimi',
        ],
        'quantity': [
            'qty', 'quantity', 'quantité', 'quantite', 'qté', 'qt', 'qte',
            'nb', 'nombre', 'count', 'amount', 'adet', 'miktar', 'menge',
            'quantita', 'quantità', 'pieces', 'pcs', 'units', 'unités',
        ],
        'price': [
            'prix', 'price', 'tarif', 'pu', 'p.u.', 'montant', 'prix ht',
            'prix unitaire', 'pu ht', 'selling', 'list price', 'net price',
            'unit price', 'sell price', 'tarif net', 'fiyat', 'prezzo',
            'preis', 'birim fiyat',
        ],
        'supplier': ['fournisseur', 'supplier', 'marque', 'brand', 'fabricant'],
        'delivery': ['délai', 'delai', 'livraison', 'delivery', 'dispo', 'disponibilité'],
        'category': ['catégorie', 'categorie', 'category', 'famille', 'type'],
        'brand': ['marque', 'brand', 'fabricant'],
    }

    @staticmethod
    def parse(file_path: str) -> List[Dict[str, Any]]:
        """Parse un fichier Excel — toutes feuilles, aucune ligne perdue."""
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl non installé. Installation: pip install openpyxl")
            return []

        products: List[Dict[str, Any]] = []
        fname = Path(file_path).name

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet_count = len(wb.sheetnames)
            logger.info(
                "event=excel_file_detected file=%s sheet_count=%d",
                fname, sheet_count,
            )
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_products = ExcelParser._parse_sheet(sheet, file_path, sheet_name)
                products.extend(sheet_products)
            wb.close()

        except Exception as exc:
            logger.error("event=excel_parse_error file=%s error=%s", fname, exc)

        logger.info(
            "event=rows_extracted file=%s total=%d",
            fname, len(products),
        )
        return products

    @staticmethod
    def _parse_sheet(sheet, file_path: str, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Parse robuste d'une feuille.
        Lit TOUTES les lignes ; supprime uniquement les lignes totalement vides.
        """
        fname = Path(file_path).name
        products: List[Dict[str, Any]] = []

        rows = list(sheet.iter_rows(values_only=True))
        rows_total_raw = len(rows)
        if not rows:
            logger.info(
                "event=sheet_empty file=%s sheet=%s", fname, sheet_name
            )
            return products

        logger.info(
            "event=rows_total_raw file=%s sheet=%s rows=%d",
            fname, sheet_name, rows_total_raw,
        )

        # ── Détection de la ligne d'en-tête ──────────────────────────────────
        # On cherche dans les 30 premières lignes la première qui contient
        # ≥2 cellules texte correspondant à des keywords connus.
        header_row_idx: Optional[int] = None
        col_indices: Dict[str, int] = {}

        for i, row in enumerate(rows[:30]):
            if not row:
                continue
            non_empty_cells = [
                str(c).strip().lower()
                for c in row
                if c is not None and str(c).strip()
            ]
            if not non_empty_cells:
                continue

            matched_fields: Dict[str, int] = {}
            for col_pos, cell_val in enumerate(
                str(c).strip().lower() if c is not None else ''
                for c in row
            ):
                if not cell_val:
                    continue
                for field, keywords in ExcelParser._COL_KEYWORDS.items():
                    if field not in matched_fields and any(kw in cell_val for kw in keywords):
                        matched_fields[field] = col_pos
                        break

            # Une ligne est un header si elle mappe ≥2 champs distincts
            if len(matched_fields) >= 2:
                header_row_idx = i
                col_indices = matched_fields
                break

        # ── Fallback : si aucun header structuré → lire à partir de la 1re ligne non-vide
        data_start: int
        has_good_mapping = len(col_indices) >= 2

        if header_row_idx is not None:
            data_start = header_row_idx + 1
        else:
            # Pas de header reconnu → données dès la première ligne non-vide
            data_start = 0
            for i, row in enumerate(rows):
                if row and any(c is not None and str(c).strip() for c in row):
                    data_start = i
                    break

        price_header = ''
        if 'price' in col_indices and header_row_idx is not None:
            hrow = rows[header_row_idx]
            price_col = col_indices['price']
            price_header = str(hrow[price_col]).strip() if price_col < len(hrow) else ''
        sheet_currency = detect_currency(price_header)

        logger.info(
            "event=excel_mapping file=%s sheet=%s header_row=%s mapped_cols=%s "
            "has_good_mapping=%s data_start=%d",
            fname, sheet_name, header_row_idx, list(col_indices.keys()),
            has_good_mapping, data_start,
        )

        # ── Parsing ligne par ligne ───────────────────────────────────────────
        rows_after_header = rows[data_start:]
        rows_skipped_empty = 0
        rows_written = 0

        for row_idx, row in enumerate(rows_after_header, start=data_start):
            # Ignorer uniquement les lignes totalement vides
            if not row or all(
                c is None or str(c).strip() == '' for c in row
            ):
                rows_skipped_empty += 1
                continue

            if has_good_mapping:
                ref, designation, qty = ExcelParser._extract_mapped(
                    row, col_indices
                )
            else:
                ref, designation, qty = ExcelParser._extract_heuristic(row)

            # Même si ref ET designation sont vides après extraction,
            # on prend le contenu brut de la ligne pour ne rien perdre.
            if not ref and not designation:
                raw = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if not raw:
                    rows_skipped_empty += 1
                    continue
                # Heuristique de dernier recours : 1re cellule = ref, 2e = désignation
                ref = raw[0] if raw else ''
                designation = raw[1] if len(raw) > 1 else raw[0]

            # Prix
            unit_price: Optional[float] = None
            row_currency = sheet_currency
            if 'price' in col_indices and col_indices['price'] < len(row):
                price_val = row[col_indices['price']]
                if isinstance(price_val, (int, float)) and price_val:
                    unit_price = float(price_val)
                    row_currency = detect_currency(str(price_val)) or sheet_currency
                elif price_val:
                    unit_price = extract_price(str(price_val))
                    row_currency = detect_currency(str(price_val)) or sheet_currency

            def _get(field: str) -> Optional[str]:
                if field in col_indices and col_indices[field] < len(row):
                    v = row[col_indices[field]]
                    return str(v).strip() if v is not None else None
                return None

            product: Dict[str, Any] = {
                'supplier_reference': ref if ref else None,
                'designation': designation if designation else None,
                'quantity': qty,
                'unit_price': unit_price,
                'supplier_name': _get('supplier'),
                'delivery_time': _get('delivery'),
                'category': _get('category'),
                'brand': _get('brand'),
                'currency': row_currency,
                'additional_data': {
                    'source_file': fname,
                    'sheet_name': sheet_name,
                    'row_index': row_idx,
                    'raw_cells': [
                        str(c).strip() if c is not None else ''
                        for c in row
                    ],
                },
            }
            products.append(product)
            rows_written += 1

        logger.info(
            "event=rows_after_filter file=%s sheet=%s "
            "rows_total_raw=%d data_rows=%d rows_skipped_empty=%d rows_extracted=%d",
            fname, sheet_name,
            rows_total_raw,
            len(rows_after_header),
            rows_skipped_empty,
            rows_written,
        )
        return products

    @staticmethod
    def _extract_mapped(
        row: tuple,
        col_indices: Dict[str, int],
    ) -> tuple:
        """Extrait (ref, designation, qty) depuis les colonnes mappées."""
        def _cell(field: str) -> Optional[str]:
            idx = col_indices.get(field)
            if idx is not None and idx < len(row):
                v = row[idx]
                return str(v).strip() if v is not None else None
            return None

        ref = _cell('reference')
        designation = _cell('designation')

        qty: Optional[int] = None
        qty_idx = col_indices.get('quantity')
        if qty_idx is not None and qty_idx < len(row):
            qty_val = row[qty_idx]
            if isinstance(qty_val, (int, float)) and qty_val > 0:
                qty = int(qty_val)
            elif qty_val:
                # Gérer "2 Adet", "3 pcs", "5 unités", etc.
                m = re.match(r'^\s*(\d+(?:[.,]\d+)?)', str(qty_val).strip())
                if m:
                    try:
                        qty = int(float(m.group(1).replace(',', '.')))
                    except ValueError:
                        pass

        return ref, designation, qty

    @staticmethod
    def _extract_heuristic(row: tuple) -> tuple:
        """
        Extraction heuristique quand aucune colonne n'est mappée.
        Parcourt les cellules de gauche à droite :
          - 1re cellule alphanumérique courte (≤40 car.) → ref
          - 1re cellule texte longue (>3 car.) non-numérique → designation
          - 1re cellule numérique entière → qty
        """
        ref: Optional[str] = None
        designation: Optional[str] = None
        qty: Optional[int] = None

        for cell in row:
            if cell is None:
                continue
            s = str(cell).strip()
            if not s:
                continue

            # Quantité : entier court (1–4 chiffres)
            if qty is None and re.match(r'^\d{1,4}$', s):
                try:
                    v = int(s)
                    if 1 <= v <= 9999:
                        qty = v
                        continue
                except ValueError:
                    pass

            # Référence : alphanumérique + tirets/slashes, longueur raisonnable
            if ref is None and len(s) <= 50 and re.match(
                r'^[A-Z0-9][A-Z0-9\-_/\.]{1,}$', s, re.IGNORECASE
            ) and any(c.isdigit() for c in s):
                ref = s
                continue

            # Désignation : texte de longueur ≥3 non purement numérique
            if designation is None and len(s) >= 3 and not s.replace('.', '').replace(',', '').isnumeric():
                designation = s

        return ref, designation, qty


class CSVParser:
    """Parser pour les fichiers CSV."""

    @staticmethod
    def parse(file_path: str) -> List[Dict[str, Any]]:
        """Parse un fichier CSV et extrait les données de tarifs."""
        import csv
        products = []

        # Essayer différents encodages et délimiteurs
        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        delimiters = [';', ',', '\t', '|']

        for encoding in encodings:
            for delimiter in delimiters:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        # Lire quelques lignes pour vérifier
                        sample = f.read(4096)
                        f.seek(0)

                        # Vérifier si le délimiteur est présent
                        if delimiter not in sample:
                            continue

                        reader = csv.DictReader(f, delimiter=delimiter)

                        for row in reader:
                            product = CSVParser._parse_row(row, file_path)
                            if product:
                                products.append(product)

                        if products:
                            return products

                except (UnicodeDecodeError, csv.Error):
                    continue
                except Exception as e:
                    logger.error(f"Erreur CSV {file_path}: {e}")
                    continue

        return products

    @staticmethod
    def _parse_row(row: Dict, file_path: str) -> Optional[Dict[str, Any]]:
        """Parse une ligne CSV."""
        # Normaliser les clés
        normalized = {k.lower().strip(): v for k, v in row.items() if k}

        # Mapping des champs
        field_mapping = {
            'reference': ['ref', 'réf', 'reference', 'référence', 'code', 'sku', 'article',
                          'item no', 'part no', 'no.', 'part number', 'item number'],
            'designation': ['designation', 'désignation', 'description', 'libellé', 'libelle',
                            'nom', 'produit', 'item description'],
            'price': ['prix', 'price', 'tarif', 'pu', 'p.u.', 'montant', 'prix ht',
                      'selling', 'list price', 'net price', 'unit price', 'sell price'],
            'supplier': ['fournisseur', 'supplier', 'marque'],
            'delivery': ['délai', 'delai', 'livraison', 'delivery']
        }

        def find_value(field_names):
            for name in field_names:
                for key in normalized:
                    if name in key:
                        return normalized[key]
            return None

        ref = find_value(field_mapping['reference'])
        designation = find_value(field_mapping['designation'])

        if not ref and not designation:
            return None

        price_str = find_value(field_mapping['price'])

        # Détecter la devise depuis la valeur du prix et/ou la clé de colonne
        price_key = next(
            (k for k in normalized if any(n in k for n in field_mapping['price'])),
            ''
        )
        currency = detect_currency(price_str or '') or detect_currency(price_key)

        return {
            'supplier_reference': ref,
            'designation': designation,
            'unit_price': extract_price(price_str) if price_str else None,
            'supplier_name': find_value(field_mapping['supplier']),
            'delivery_time': find_value(field_mapping['delivery']),
            'currency': currency,
            'additional_data': {'source_file': Path(file_path).name}
        }


class ImageParser:
    """Parser pour les images (via OCR)."""

    @staticmethod
    def parse(file_path: str) -> List[Dict[str, Any]]:
        """Parse une image via OCR et extrait les données."""
        products = []

        try:
            import pytesseract
            from PIL import Image
        except ImportError:
            logger.warning("pytesseract ou Pillow non installé. Installation: pip install pytesseract pillow")
            return []

        try:
            image = Image.open(file_path)
            text = pytesseract.image_to_string(image, lang='fra+eng')

            # Extraction heuristique du texte OCR
            products = PDFParser._extract_products_from_text(text, file_path)

        except Exception as e:
            logger.error(f"Erreur OCR image {file_path}: {e}")

        return products


def parse_file(file_path: str) -> List[Dict[str, Any]]:
    """Parse un fichier selon son type et retourne les produits extraits."""
    file_type = get_file_type(file_path)

    if not file_type:
        logger.warning(f"Type de fichier non supporté: {file_path}")
        return []

    parsers = {
        'pdf': PDFParser.parse,
        'excel': ExcelParser.parse,
        'csv': CSVParser.parse,
        'image': ImageParser.parse
    }

    parser = parsers.get(file_type)
    if parser:
        logger.info(f"Parsing {file_type}: {file_path}")
        return parser(file_path)

    return []


def scan_folder(folder_path: str, recursive: bool = True) -> List[Dict[str, Any]]:
    """Scanne un dossier et retourne la liste des fichiers supportés."""
    folder = Path(folder_path)

    if not folder.exists():
        logger.error(f"Dossier non trouvé: {folder_path}")
        return []

    files = []

    if recursive:
        for ext in ALL_SUPPORTED_EXTENSIONS:
            files.extend(folder.rglob(f"*{ext}"))
    else:
        for ext in ALL_SUPPORTED_EXTENSIONS:
            files.extend(folder.glob(f"*{ext}"))

    result = []
    for file_path in files:
        stat = file_path.stat()
        result.append({
            'path': str(file_path),
            'name': file_path.name,
            'type': get_file_type(str(file_path)),
            'size': stat.st_size,
            'modified': datetime.fromtimestamp(stat.st_mtime)
        })

    return result
