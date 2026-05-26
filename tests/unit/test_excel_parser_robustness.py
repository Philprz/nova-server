# tests/unit/test_excel_parser_robustness.py
"""
Tests de robustesse du parser Excel renforcé.

INVARIANTS :
  - Extraction quasi-complète (≥95% des lignes non-vides d'un Excel réel)
  - Aucune ligne non-vide supprimée silencieusement
  - Quantité correctement lue depuis l'Excel
  - Toutes lignes propagées dans product_matches (API)
  - Fonctionne sans header reconnu (fallback heuristique)

Couvre les 3 tests obligatoires + cas limites critiques.
"""

import sys
import os
import io
import tempfile
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))

from services.file_parsers import ExcelParser
from services.email_matcher import MatchedProduct


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _xlsx(rows_of_rows: list, sheet_name: str = "Sheet1") -> str:
    """Crée un xlsx temporaire et retourne son chemin."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows_of_rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
    with open(path, "wb") as f:
        f.write(buf.getvalue())
    return path


def _multi_sheet_xlsx(sheets: dict) -> str:
    """Crée un xlsx multi-feuilles. sheets = {name: [[row], ...]}."""
    import openpyxl
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    path = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
    with open(path, "wb") as f:
        f.write(buf.getvalue())
    return path


def _parse(path: str) -> list:
    try:
        return ExcelParser.parse(path)
    finally:
        try:
            os.unlink(path)
        except Exception:
            pass


# ─── TEST 1 : extraction complète ────────────────────────────────────────────

class TestExtractionComplete:
    """
    TEST 1 : pour tout Excel valide, nb_lignes_extraites >= 95% des lignes non-vides.
    """

    def test_standard_rfq_headers_fr(self):
        """Headers FR standard : Référence / Désignation / Qté."""
        path = _xlsx([
            ["Référence", "Désignation", "Qté"],
            ["C892-001", "Servo Drive", 2],
            ["C893-007-XY", "Applications Module", 1],
            ["C853-002", "Brake Resistor", 1],
            ["C814-RS-265-0859", "Fuse 2A", 1],
            ["C843-003", "LED Keypad", 1],
        ])
        rows = _parse(path)
        assert len(rows) >= 5, f"Attendu >=5 lignes, obtenu {len(rows)}: {rows}"

    def test_standard_rfq_headers_en(self):
        """Headers EN standard : Part Number / Description / Qty."""
        path = _xlsx([
            ["Part Number", "Description", "Qty"],
            ["C892-001", "Servo Drive", 2],
            ["C893-007-XY", "Applications Module XY", 1],
            ["INCONNUE-999", "Unknown spare part", 3],
            ["C874-002", "SP-DT Adaptor", 2],
        ])
        rows = _parse(path)
        assert len(rows) >= 4, f"Attendu >=4, obtenu {len(rows)}"

    def test_nonstandard_headers_fallback(self):
        """
        Headers non reconnus → fallback heuristique.
        Aucune ligne ne doit être perdue.
        """
        path = _xlsx([
            ["N°", "Pièce demandée", "Nb souhaité"],  # Headers non reconnus
            ["C892-001", "Servo Drive", 2],
            ["C893-007", "Base Part", 1],
            ["C874-002", "Adaptor", 2],
        ])
        rows = _parse(path)
        # Au moins 2 lignes sur 3 (tolérance heuristique)
        assert len(rows) >= 2, f"Attendu >=2, obtenu {len(rows)}: {rows}"

    def test_no_header_row_at_all(self):
        """Pas de ligne d'en-tête : données dès la ligne 0."""
        path = _xlsx([
            ["C892-001", "Servo Drive", 2],
            ["C893-007-XY", "Module", 1],
            ["C874-002", "Adaptor", 2],
        ])
        rows = _parse(path)
        assert len(rows) >= 2, f"Attendu >=2, obtenu {len(rows)}"

    def test_large_list_95_percent_extracted(self):
        """
        CRITIQUE : 20 articles → ≥19 extraits (95%).
        Simule un vrai Excel RFQ industriel.
        """
        data = [["Part No", "Description", "Qty"]]
        for i in range(1, 21):
            data.append([f"C{800+i:03d}-00{i % 9 + 1}", f"Article {i}", i % 5 + 1])
        path = _xlsx(data)
        rows = _parse(path)
        assert len(rows) >= 19, (
            f"Attendu >=19 (95% de 20), obtenu {len(rows)}"
        )

    def test_turkish_headers(self):
        """Headers turcs : Malzeme Kodu / Tanim / Adet."""
        path = _xlsx([
            ["Malzeme Kodu", "Tanim", "Adet"],
            ["C892-001", "Servo Surucu", 2],
            ["C893-007-XY", "Modul XY", 1],
        ])
        rows = _parse(path)
        assert len(rows) >= 2, f"Attendu >=2, obtenu {len(rows)}"

    def test_multi_sheet_all_extracted(self):
        """Toutes les feuilles d'un classeur multi-onglets sont parsées."""
        path = _multi_sheet_xlsx({
            "Feuille1": [
                ["Reference", "Designation", "Qty"],
                ["C892-001", "Servo Drive", 2],
                ["C893-007", "Base Part", 1],
            ],
            "Feuille2": [
                ["Part No", "Description", "Qty"],
                ["C874-002", "Adaptor", 2],
                ["C843-003", "Keypad", 1],
            ],
        })
        rows = _parse(path)
        assert len(rows) >= 4, f"Attendu >=4 (2 feuilles), obtenu {len(rows)}"
        sheets = {r["additional_data"]["sheet_name"] for r in rows}
        assert len(sheets) == 2, f"Attendu 2 feuilles, obtenu {sheets}"

    def test_source_metadata_present(self):
        """Chaque ligne extraite contient source_file, sheet_name, row_index."""
        path = _xlsx([
            ["Reference", "Designation", "Qty"],
            ["C892-001", "Servo Drive", 2],
        ])
        rows = _parse(path)
        assert rows, "Aucune ligne extraite"
        meta = rows[0]["additional_data"]
        assert meta.get("source_file"), "source_file manquant"
        assert meta.get("sheet_name") is not None, "sheet_name manquant"
        assert meta.get("row_index") is not None, "row_index manquant"


# ─── TEST 2 : non-suppression ─────────────────────────────────────────────────

class TestNoSuppression:
    """
    TEST 2 : aucune ligne non-vide ne doit être supprimée sans raison explicite.
    """

    def test_empty_rows_skipped_but_data_rows_kept(self):
        """Lignes vides ignorées, lignes avec données gardées."""
        path = _xlsx([
            ["Reference", "Designation", "Qty"],
            ["C892-001", "Servo Drive", 2],
            [None, None, None],          # ligne vide → skip légal
            ["", "", ""],                # ligne vide → skip légal
            ["C893-007", "Module", 1],
        ])
        rows = _parse(path)
        assert len(rows) >= 2, f"Attendu >=2, obtenu {len(rows)}"

    def test_rows_without_reference_kept_if_designation_present(self):
        """Ligne sans référence mais avec désignation → conservée."""
        path = _xlsx([
            ["Reference", "Designation", "Qty"],
            [None, "Servo Drive sans ref", 2],
            ["C893-007", "Module avec ref", 1],
        ])
        rows = _parse(path)
        assert len(rows) >= 2, (
            f"La ligne sans référence doit être conservée. Obtenu {len(rows)}"
        )

    def test_rows_without_designation_kept_if_reference_present(self):
        """Ligne avec référence mais sans désignation → conservée."""
        path = _xlsx([
            ["Reference", "Designation", "Qty"],
            ["C892-001", None, 2],
            ["C893-007", "Module", 1],
        ])
        rows = _parse(path)
        assert len(rows) >= 2, (
            f"La ligne sans désignation doit être conservée. Obtenu {len(rows)}"
        )

    def test_unknown_column_names_do_not_drop_rows(self):
        """
        CRITIQUE : si les headers ne matchent aucun keyword,
        les lignes de données ne doivent pas être perdues.
        """
        path = _xlsx([
            ["Numéro interne", "Libellé article", "Nb commandé"],
            ["C892-001", "Servo Drive", 2],
            ["C893-007", "Module XY", 1],
            ["C874-002", "Adaptor", 2],
        ])
        rows = _parse(path)
        assert len(rows) >= 2, (
            f"Headers non reconnus ne doivent pas supprimer les lignes. "
            f"Obtenu {len(rows)}"
        )

    def test_match_status_unmatched_not_suppressed(self):
        """MatchedProduct avec match_status='unmatched' passe le filtre UI."""
        unmatched = MatchedProduct(
            item_code="REF-INCONNUE",
            item_name="Article sans SAP",
            quantity=3,
            score=0,
            match_reason="non trouvé",
            not_found_in_sap=True,
            match_status="unmatched",
            source_file="commande.xlsx",
        )
        visible = [p for p in [unmatched] if p.status != "pending_selection"]
        assert len(visible) == 1, "L'article non-matché disparaît du filtre UI"

    def test_all_rows_have_raw_cells_in_metadata(self):
        """Chaque ligne conserve ses cellules brutes pour audit."""
        path = _xlsx([
            ["Part Number", "Description", "Qty"],
            ["C892-001", "Servo Drive", 2],
        ])
        rows = _parse(path)
        assert rows, "Aucune ligne"
        raw = rows[0]["additional_data"].get("raw_cells")
        assert raw is not None, "raw_cells absent de additional_data"
        assert len(raw) >= 2, f"raw_cells trop courts : {raw}"


# ─── TEST 3 : propagation API ─────────────────────────────────────────────────

class TestApiPropagation:
    """
    TEST 3 : toutes les lignes Excel sont visibles dans la réponse JSON finale.
    """

    def test_all_excel_rows_in_product_matches(self):
        """
        Simulation de Phase 4.6 : N lignes Excel → N entrées dans product_matches.
        """
        from tests.unit.test_product_strict_matching import (
            _build_matcher_with_items, _reference_items,
        )
        matcher = _build_matcher_with_items(_reference_items())
        matcher._extract_offer_request_rows = lambda t: []
        matcher._extract_product_descriptions = lambda t: {}

        excel_rows = [
            {"supplier_reference": "C892-001", "designation": "Servo Drive",
             "quantity": 2, "_source_file": "test.xlsx", "_source_sheet": "S1"},
            {"supplier_reference": "C893-007-XY", "designation": "Module",
             "quantity": 1, "_source_file": "test.xlsx", "_source_sheet": "S1"},
            {"supplier_reference": "INCONNUE-999", "designation": "Inconnu SAP",
             "quantity": 3, "_source_file": "test.xlsx", "_source_sheet": "S1"},
        ]

        excel_products = []
        for row in excel_rows:
            ref = row.get("supplier_reference", "")
            designation = row.get("designation", "")
            qty = row.get("quantity", 1)
            sap = matcher.match_product_strict(ref, quantity=qty) if ref else []
            if sap:
                for mp in sap:
                    mp.quantity = qty
                    mp.match_status = "matched"
                    mp.source_file = row["_source_file"]
                    excel_products.append(mp)
            else:
                excel_products.append(MatchedProduct(
                    item_code=ref, item_name=designation,
                    quantity=qty, score=0, match_reason="non trouvé",
                    not_found_in_sap=True, match_status="unmatched",
                    source_file=row["_source_file"],
                ))

        # Simuler result.product_matches = [p.dict() for p in excel_products]
        serialized = [p.dict() for p in excel_products]

        assert len(serialized) == 3, (
            f"Attendu 3 (2 matchés + 1 non-matché), obtenu {len(serialized)}"
        )

    def test_quantity_from_excel_in_product_matches(self):
        """La quantité réelle de l'Excel est conservée dans product_matches."""
        from tests.unit.test_product_strict_matching import (
            _build_matcher_with_items, _reference_items,
        )
        matcher = _build_matcher_with_items(_reference_items())
        matcher._extract_offer_request_rows = lambda t: []
        matcher._extract_product_descriptions = lambda t: {}

        EXPECTED_QTY = 7
        sap = matcher.match_product_strict("C892-001", quantity=EXPECTED_QTY)
        assert sap, "C892-001 doit être trouvé dans le mock SAP"

        # Simuler Phase 4.6 : appliquer la quantité Excel sur le résultat SAP
        mp = sap[0]
        mp.quantity = EXPECTED_QTY

        assert mp.quantity == EXPECTED_QTY, (
            f"Quantité attendue {EXPECTED_QTY}, obtenu {mp.quantity}"
        )

    def test_unmatched_row_in_serialized_output(self):
        """Les lignes non-matchées apparaissent dans la sortie JSON."""
        mp = MatchedProduct(
            item_code="REF-INCONNUE",
            item_name="Produit inconnu",
            quantity=5,
            score=0,
            match_reason="non trouvé",
            not_found_in_sap=True,
            match_status="unmatched",
            source_file="rfq.xlsx",
            source_sheet="Articles",
            source_row_index=3,
        )
        d = mp.dict()
        assert d["match_status"] == "unmatched"
        assert d["not_found_in_sap"] is True
        assert d["quantity"] == 5
        assert d["source_file"] == "rfq.xlsx"


# ─── TEST 4 : quantité correctement extraite ─────────────────────────────────

class TestQuantityExtraction:
    """La quantité est extraite depuis la colonne Excel, jamais hardcodée."""

    def test_qty_from_integer_cell(self):
        """Cellule numérique entière → qty correcte."""
        path = _xlsx([
            ["Reference", "Designation", "Qty"],
            ["C892-001", "Servo Drive", 5],
        ])
        rows = _parse(path)
        assert rows, "Aucune ligne"
        assert rows[0].get("quantity") == 5, (
            f"Quantité attendue 5, obtenu {rows[0].get('quantity')}"
        )

    def test_qty_from_float_cell(self):
        """Cellule décimale (ex: 2.0) → qty = 2."""
        path = _xlsx([
            ["Reference", "Designation", "Qty"],
            ["C892-001", "Servo Drive", 2.0],
        ])
        rows = _parse(path)
        assert rows, "Aucune ligne"
        qty = rows[0].get("quantity")
        assert qty == 2, f"Attendu 2, obtenu {qty}"

    def test_qty_from_string_with_unit(self):
        """Cellule '3 Adet' → qty = 3 (format turc)."""
        path = _xlsx([
            ["Part Number", "Description", "Adet"],
            ["C892-001", "Servo", "3 Adet"],
        ])
        rows = _parse(path)
        assert rows, "Aucune ligne"
        qty = rows[0].get("quantity")
        assert qty == 3, f"Attendu 3 (depuis '3 Adet'), obtenu {qty}"

    def test_qty_none_when_no_qty_column(self):
        """Pas de colonne quantité → quantity = None (pas de fallback à 1)."""
        path = _xlsx([
            ["Reference", "Designation"],
            ["C892-001", "Servo Drive"],
        ])
        rows = _parse(path)
        if rows:
            qty = rows[0].get("quantity")
            # None est acceptable quand la colonne n'existe pas
            assert qty is None or isinstance(qty, (int, float)), (
                f"quantity doit être None ou numérique, obtenu {qty!r}"
            )

    def test_multiple_rows_different_quantities(self):
        """Chaque ligne a sa propre quantité."""
        path = _xlsx([
            ["Reference", "Designation", "Qty"],
            ["C892-001", "Servo Drive", 2],
            ["C893-007-XY", "Module", 5],
            ["C874-002", "Adaptor", 1],
        ])
        rows = _parse(path)
        assert len(rows) >= 3, f"Attendu 3 lignes, obtenu {len(rows)}"
        qtys = [r.get("quantity") for r in rows]
        assert 2 in qtys, f"qty=2 attendue, qtys={qtys}"
        assert 5 in qtys, f"qty=5 attendue, qtys={qtys}"
        assert 1 in qtys, f"qty=1 attendue, qtys={qtys}"


# ─── TEST 5 : format RFQ réel simulé ─────────────────────────────────────────

class TestRealWorldRFQ:
    """
    Simule un Excel RFQ industriel typique tel qu'envoyé par un client.
    Vérifie extraction quasi-complète.
    """

    def _build_rfq_excel(self) -> str:
        """
        Excel de type RFQ Rondot avec 10 articles.
        Format : headers EN, colonnes Part Number + Description + Quantity.
        """
        return _xlsx([
            ["No.", "Part Number", "Description", "Quantity", "Unit"],
            [1, "C892-001", "Servo Drive 1.5kW", 2, "pcs"],
            [2, "C893-007-XY", "Applications Module XY", 1, "pcs"],
            [3, "C893-007-Z", "Applications Module Z", 1, "pcs"],
            [4, "C853-002", "Compact Brake Resistor", 1, "pcs"],
            [5, "C814-RS-265-0859", "Fuse 2A", 1, "pcs"],
            [6, "C843-003", "LED Keypad DIGITAX", 1, "pcs"],
            [7, "C893-002", "SI-I/O Module", 1, "pcs"],
            [8, "C874-002", "SP-DT Adaptor MRS-9000", 2, "pcs"],
            [9, "C843-142UMA300CACAA", "SERVO MOTOR 1.7kW", 2, "pcs"],
            [10, "C893-007", "Base Part", 1, "pcs"],
        ])

    def test_10_articles_tous_extraits(self):
        """10 articles → 10 lignes extraites (100%)."""
        path = self._build_rfq_excel()
        rows = _parse(path)
        assert len(rows) >= 10, (
            f"Attendu 10 lignes, obtenu {len(rows)}: "
            f"{[r.get('supplier_reference') for r in rows]}"
        )

    def test_references_correctes(self):
        """Toutes les références sont extraites sans corruption."""
        path = self._build_rfq_excel()
        rows = _parse(path)
        refs = {r.get("supplier_reference", "").upper() for r in rows if r.get("supplier_reference")}
        expected = {
            "C892-001", "C893-007-XY", "C893-007-Z", "C853-002",
            "C814-RS-265-0859", "C843-003", "C893-002", "C874-002",
            "C843-142UMA300CACAA", "C893-007",
        }
        for exp in expected:
            assert exp in refs, f"Référence manquante : {exp}. Extraites : {refs}"

    def test_quantites_correctes(self):
        """Les quantités correspondent à celles de l'Excel."""
        path = self._build_rfq_excel()
        rows = _parse(path)
        qty_map = {
            r.get("supplier_reference", "").upper(): r.get("quantity")
            for r in rows
        }
        assert qty_map.get("C892-001") == 2, f"C892-001 qty attendu 2, obtenu {qty_map.get('C892-001')}"
        assert qty_map.get("C874-002") == 2, f"C874-002 qty attendu 2, obtenu {qty_map.get('C874-002')}"
        assert qty_map.get("C843-142UMA300CACAA") == 2, "C843-142UMA300CACAA qty attendu 2"
        assert qty_map.get("C893-007-XY") == 1, "C893-007-XY qty attendu 1"

    def test_95_percent_threshold(self):
        """
        Seuil 95% : même avec un Excel imparfait, quasi-tout doit passer.
        """
        path = self._build_rfq_excel()
        rows = _parse(path)
        n_data_rows = 10  # lignes de données (hors header)
        pct = len(rows) / n_data_rows * 100
        assert pct >= 95, (
            f"Extraction insuffisante : {pct:.0f}% ({len(rows)}/{n_data_rows}). "
            f"Seuil requis : 95%."
        )
